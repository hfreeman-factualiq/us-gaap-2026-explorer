#!/usr/bin/env python3
"""Build per-company ontology trees from the base template tree, pruned to populateable concepts.

Rule: if A = B + C and the company cannot populate B and C (and does not
report A directly), omit A, B, and C from that company's tree.

Pipeline:
  1. Extract line-item labels from financial statements and strategic/operational
     evidence from board books and investor memos
  2. Map labels → ontology concepts via exact matching + Claude (LLM)
  3. Gap analysis via Claude: important unmapped lines become new company nodes
     attached under the right parent with relationships/formulas
  4. Close under formulas when ALL inputs exist
  5. Prune the base browse forest to kept nodes + structural ancestors
  6. Emit company_trees/{slug}/taxonomy-data.json (+ viewer copies)
"""

from __future__ import annotations

import argparse
import hashlib
import json
import logging
import os
import re
import shutil
import sys
import threading
import time
from collections import Counter, defaultdict
from concurrent.futures import ProcessPoolExecutor, ThreadPoolExecutor, as_completed
from datetime import datetime, timezone
from pathlib import Path

import yaml
from openpyxl import load_workbook

import doc_extract
from llm_mapping import (
    build_candidate_list,
    map_labels_with_llm,
    propose_gaps_with_llm,
)

ROOT = Path(__file__).resolve().parent
_BASE_SNAPSHOTS = [
    ROOT / "financial_ontology_2026_v1.31" / "taxonomy-data.json",
    ROOT / "financial_ontology_2026_v1.3" / "taxonomy-data.json",
    ROOT / "financial_ontology_2026_v1.2" / "taxonomy-data.json",
]
ONTOLOGY_PATH = next((p for p in _BASE_SNAPSHOTS if p.exists()), ROOT / "taxonomy-data.json")
SYNONYM_PATH = ROOT / "mappings" / "label_synonyms.yaml"
DOCUMENT_PACK_DIR = ROOT / "company_document_pack"
OUT_DIR = ROOT / "company_trees"
VIEWER_DIR = ROOT / "viewer" / "companies"
PAGES_DIR = ROOT / "companies"
LOG_DIR = ROOT / "logs"
PROGRESS_PATH = OUT_DIR / "build_progress.json"

log = logging.getLogger("build")

COMPANIES = [
    {
        "slug": "orijin",
        "name": "Orijin",
        "folder": "Orijin",
        "prefer_glob": ["*Consolidated*", "*December*", "*Dec*"],
        "sheet_keywords": {
            "is": ["p&l", "income", "vs.", "ytd", "ordinary"],
            "bs": ["balance"],
            "cf": ["cash flow", "cash flows"],
        },
        "skip_sheets": ["ar aging", "ap aging", "cash forecast", "by month"],
    },
    {
        "slug": "climb-credit",
        "name": "Climb Credit",
        "folder": "Climb",
        "prefer_glob": ["*Consolidated*Dec*", "*Consolidated*"],
        "sheet_keywords": {
            "is": ["profit", "loss", "income", "p&l"],
            "bs": ["balance"],
            "cf": ["cash flow"],
        },
        "skip_sheets": [],
    },
    {
        "slug": "mantra-health",
        "name": "Mantra Health",
        "folder": "Mantra",
        "prefer_glob": ["*Dec*", "*YTD*"],
        "sheet_keywords": {
            "is": ["p&l", "income", "profit"],
            "bs": ["bs", "balance"],
            "cf": ["cf", "cash"],
        },
        "skip_sheets": [],
    },
    {
        "slug": "brains-and-motion",
        "name": "Brains and Motion",
        "folder": "BAM",
        "prefer_glob": ["*Dec*", "*Financial*"],
        "sheet_keywords": {
        "is": ["summary projections", "summary", "category p&l", "consolidated dept p&l"],
        "bs": ["balance sheet"],
        "cf": ["cash flow"],
      },
      "skip_sheets": [
        "documentation",
        "graph",
        "payroll",
        "pipeline",
        "bookings",
        "unit econ",
        "assumptions",
        "mapping",
        "instructor",
        "zero check",
        "msa",
        "bva",
        "location",
        "prepaid schedule",
        "actual revenue",
        "active contracts",
        "university model",
        "db_act",
        "revenue summary",
        "revenue & cor",
        "metrics",
        "actpl",
        "dept p&l - main",
        "dept p&l (by category)",
        "kidz",
        "brains and motion edu",
        "consol pl",
        "actual",
        "budget",
      ],
    },
    {
        "slug": "knack",
        "name": "Knack",
        "folder": "Knack",
        "prefer_glob": ["*Financial Model*", "*Investor*Model*", "*2025*", "*2026*"],
        "sheet_keywords": {
            "is": ["income", "p&l", "profit", "loss", "financial model"],
            "bs": ["balance"],
            "cf": ["cash flow", "cashflow"],
        },
        "skip_sheets": ["assumptions", "cover", "instructions"],
    },
]

# Statement buckets used to hang unmatched-but-mapped detail under extensions.
ASSET_CONCEPTS = {
    "Assets",
    "AssetsCurrent",
    "AssetsNoncurrent",
    "CashAndCashEquivalentsAtCarryingValue",
    "ReceivablesNetCurrent",
    "InventoryNet",
    "PrepaidExpenseCurrent",
    "PropertyPlantAndEquipmentNet",
    "IntangibleAssetsNetIncludingGoodwill",
    "OperatingLeaseRightOfUseAsset",
    "OtherAssetsNoncurrent",
}
LIAB_CONCEPTS = {
    "Liabilities",
    "LiabilitiesCurrent",
    "LiabilitiesNoncurrent",
    "AccountsPayableAndAccruedLiabilitiesCurrent",
    "ContractWithCustomerLiability",
    "ContractWithCustomerLiabilityCurrent",
    "ShortTermBorrowings",
    "LongTermDebt",
    "OperatingLeaseLiability",
    "OtherLiabilitiesCurrent",
    "OtherLiabilitiesNoncurrent",
    "DeferredIncomeTaxLiabilitiesNet",
}
EQUITY_CONCEPTS = {
    "StockholdersEquity",
    "RetainedEarningsAccumulatedDeficit",
    "AccumulatedOtherComprehensiveIncomeLossNetOfTax",
    "MinorityInterest",
    "CommonStockValue",
    "PreferredStockValue",
    "AdditionalPaidInCapital",
    "TreasuryStockValue",
}
IS_CONCEPTS = {
    "Revenues",
    "CostOfRevenue",
    "GrossProfit",
    "OperatingIncomeLoss",
    "OperatingExpenses",
    "ResearchAndDevelopmentExpense",
    "SellingGeneralAndAdministrativeExpense",
    "InterestAndDebtExpense",
    "IncomeTaxExpenseBenefit",
    "IncomeLossFromContinuingOperationsBeforeIncomeTaxesExtraordinaryItemsNoncontrollingInterest",
    "ProfitLoss",
    "EarningsPerShareDiluted",
    "EarningsPerShareBasic",
    "NonoperatingIncomeExpense",
    "DepreciationDepletionAndAmortization",
    "nongaap:EBITDA",
}
CF_CONCEPTS = {
    "NetCashProvidedByUsedInOperatingActivities",
    "NetCashProvidedByUsedInInvestingActivities",
    "NetCashProvidedByUsedInFinancingActivities",
    "CashCashEquivalentsRestrictedCashAndRestrictedCashEquivalentsPeriodIncreaseDecreaseExcludingExchangeRateEffect",
    "PaymentsToAcquirePropertyPlantAndEquipment",
    "ProfitLoss",
    "AdjustmentsToReconcileNetIncomeLossToCashProvidedByUsedInOperatingActivities",
}

# Explicit formulas used when browse tc is empty / noisy (ALL inputs required).
EXPLICIT_FORMULAS: dict[str, list[str]] = {
    "GrossProfit": ["Revenues", "CostOfRevenue"],
    "OperatingIncomeLoss": ["GrossProfit", "OperatingExpenses"],
    "CashCashEquivalentsRestrictedCashAndRestrictedCashEquivalentsPeriodIncreaseDecreaseExcludingExchangeRateEffect": [
        "NetCashProvidedByUsedInOperatingActivities",
        "NetCashProvidedByUsedInInvestingActivities",
        "NetCashProvidedByUsedInFinancingActivities",
    ],
    "nongaap:WorkingCapital": ["AssetsCurrent", "LiabilitiesCurrent"],
    "Assets": ["AssetsCurrent", "AssetsNoncurrent"],
    "Liabilities": ["LiabilitiesCurrent", "LiabilitiesNoncurrent"],
}

SKIP_LABEL_RE = re.compile(
    r"^(draft|confidential|amounts in|parameter|zero check|common sized|"
    r"section [a-z]|customer:|current month|reporting book|as of date|"
    r"location:|bookings|choose method|remarks|variance|forecast|budget|"
    r"actual|q[1-4]|jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec|"
    r"ytd|fy\s*\d|ye\s*dec|\d{4}$)",
    re.I,
)
GL_CODE_RE = re.compile(r"^\d{3,5}\s*[-–]")
NOISE_RE = re.compile(r"^[\W_]+$")


def canon_label(text: str) -> str:
    s = str(text).lower().strip()
    s = s.replace("&", " and ").replace("/", " ")
    s = re.sub(r"\(.*?\)", " ", s)
    s = re.sub(r"[^a-z0-9\s]+", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    # Drop leading total-/net- noise variants already covered by synonyms
    return s


def is_numbery(val) -> bool:
    if isinstance(val, (int, float)) and not isinstance(val, bool):
        return True
    if isinstance(val, str):
        t = val.strip().replace(",", "").replace("$", "").replace("(", "-").replace(")", "")
        if not t or t in {"-", "—", "–"}:
            return False
        try:
            float(t)
            return True
        except ValueError:
            return False
    return False


def load_synonyms() -> tuple[dict[str, str], dict[str, list[str]]]:
    raw = yaml.safe_load(SYNONYM_PATH.read_text(encoding="utf-8"))
    syn = {canon_label(k): v for k, v in (raw.get("synonyms") or {}).items()}
    aliases = raw.get("input_aliases") or {}
    return syn, aliases


def list_company_files(cfg: dict) -> list[Path]:
    folder = DOCUMENT_PACK_DIR / cfg["folder"] / "01_financial_statements"
    files = []
    for pat in ("*.xlsx", "*.xls", "*.xlsb"):
        files.extend(folder.glob(pat))
    files = [f for f in files if not f.name.startswith("~$")]
    name_filter = cfg.get("name_filter")
    if name_filter:
        filtered = [f for f in files if name_filter(f)]
        if filtered:
            files = filtered
    # Prefer latest by YYYY_MM prefix when present
    def sort_key(p: Path):
        m = re.match(r"(\d{4})_(\d{2})", p.name)
        if m:
            return (int(m.group(1)), int(m.group(2)), p.name)
        return (0, 0, p.name)

    files.sort(key=sort_key, reverse=True)
    return files


def pick_source_files(cfg: dict, files: list[Path], limit: int | None = None) -> list[Path]:
    if not files:
        return []
    if limit is None:
        # The document pack is curated input: process every financial workbook.
        return files
    preferred = []
    for f in files:
        low = f.name.lower()
        if any(re.search(g.replace("*", ".*"), low) for g in cfg.get("prefer_glob") or []):
            preferred.append(f)
    chosen = preferred[:limit] if preferred else files[:limit]
    # Always include the chronologically latest file
    if files[0] not in chosen:
        chosen = [files[0]] + chosen
    # de-dupe preserve order
    out, seen = [], set()
    for f in chosen:
        if f not in seen:
            out.append(f)
            seen.add(f)
    return out[:limit]


NARRATIVE_SUFFIXES = doc_extract.NARRATIVE_SUFFIXES
NARRATIVE_KEYWORDS_RE = re.compile(
    r"\b("
    r"arr|mrr|revenue|gross margin|ebitda|burn|runway|cash|forecast|budget|"
    r"pipeline|booking|retention|churn|customer|logo|cac|ltv|sales cycle|"
    r"headcount|hiring|vacanc|attrition|turnover|leadership|reorg|"
    r"strategy|strategic|vision|mission|goal|objective|tactic|policy|"
    r"decision|rule|capabilit|value stream|value proposition|"
    r"supply chain|source|deliver|reliability|agility|"
    r"deployment|lead time|failure rate|restore service|mttr|incident|problem|sla|"
    r"risk|threat|vulnerab|loss magnitude|governance|control|audit|"
    r"esg|sustainab|climate|emission|impact|outcome|underserved"
    r")\b",
    re.I,
)


def list_narrative_files(cfg: dict) -> list[Path]:
    folder = DOCUMENT_PACK_DIR / cfg["folder"]
    files = [
        p
        for p in folder.rglob("*")
        if p.is_file()
        and p.suffix.lower() in NARRATIVE_SUFFIXES
        and not p.name.startswith("~$")
        and "01_financial_statements" not in {part.lower() for part in p.parts}
    ]
    return sorted(files, key=lambda p: str(p).lower())


extract_narrative_text = doc_extract.extract_narrative_text


def clean_narrative_line(raw: str) -> str:
    line = re.sub(r"\s+", " ", raw).strip()
    line = re.sub(r"^[\u2022\u25cf\u25aa\u25e6\-–—\s]+", "", line)
    return line.strip()


MAX_PHRASE_WORDS = 6


def count_phrase_hits(words: list[str], phrases: set[str]) -> Counter:
    """Count occurrences of target phrases with one sliding pass per phrase length.

    Scanning the corpus once per phrase (thousands of regex passes over megabytes
    of text) was the slowest step of the build; bucketing by length makes it a
    handful of passes with dict lookups.
    """
    by_len: dict[int, set[str]] = defaultdict(set)
    for phrase in phrases:
        n = phrase.count(" ") + 1
        if n <= MAX_PHRASE_WORDS:
            by_len[n].add(phrase)
    hits: Counter = Counter()
    total = len(words)
    for size, targets in by_len.items():
        if size == 1:
            for word in words:
                if word in targets:
                    hits[word] += 1
            continue
        for i in range(total - size + 1):
            phrase = " ".join(words[i : i + size])
            if phrase in targets:
                hits[phrase] += 1
    return hits


def extract_narrative_evidence(
    cfg: dict,
    concepts: dict,
    *,
    texts: dict[str, str],
    sources: list[Path] | None = None,
    max_labels: int = 250,
) -> tuple[list[Path], list[str], dict[str, str]]:
    """Return sources, LLM evidence lines, and exact concept mentions.

    Document text is extracted (and cached) up front by ``extract_all_documents``;
    this step only mines the text.
    """
    sources = sources if sources is not None else list_narrative_files(cfg)
    labels: list[str] = []
    corpus_words: list[str] = []

    for path in sources:
        text = texts.get(str(path)) or ""
        if not text:
            continue
        corpus_words.extend(re.findall(r"[a-z0-9]+", text.lower()))
        for raw in text.splitlines():
            line = clean_narrative_line(raw)
            if not (3 <= len(line) <= 140):
                continue
            if NARRATIVE_KEYWORDS_RE.search(line):
                labels.append(line)

    # Unique, stable, bounded evidence lines for LLM mapping and gap analysis.
    unique_labels: list[str] = []
    seen_labels: set[str] = set()
    for label in labels:
        key = canon_label(label)
        if not key or key in seen_labels:
            continue
        seen_labels.add(key)
        unique_labels.append(label)
        if len(unique_labels) >= max_labels:
            break

    targets: dict[str, tuple[str, str]] = {}
    for cid, concept in concepts.items():
        if not is_mappable_concept(cid, concept):
            continue
        if concept.get("stdKind") not in {"signal", "standardClass"} and concept.get("layer") != "nongaap":
            continue
        if concept.get("category") == "scraped-class":
            continue
        label = clean_narrative_line(str(concept.get("l") or ""))
        if not label:
            continue
        words = re.findall(r"[a-z0-9]+", label.lower())
        if not words or len(words) > MAX_PHRASE_WORDS:
            continue
        # Avoid broad false positives such as Plan, Risk, Make, Goal.
        if len(words) == 1 and len(words[0]) < 8 and words[0] not in {"arr", "mrr", "nrr", "ebitda", "mttr", "cac"}:
            continue
        targets[cid] = (label, " ".join(words))

    hits = count_phrase_hits(corpus_words, {phrase for _, phrase in targets.values()})
    exact: dict[str, str] = {}
    for cid, (label, phrase) in targets.items():
        count = hits.get(phrase, 0)
        if count:
            exact[cid] = f"{label} (document mention; {count} occurrence{'s' if count != 1 else ''})"

    return sources, unique_labels, exact


def classify_sheet(name: str, cfg: dict) -> str | None:
    low = name.lower().strip()
    for skip in cfg.get("skip_sheets") or []:
        if skip.lower() in low:
            return None
    for kind, keys in (cfg.get("sheet_keywords") or {}).items():
        if any(k in low for k in keys):
            return kind
    return None


def iter_sheet_rows_xlsx(path: Path):
    wb = load_workbook(path, data_only=True, read_only=True)
    try:
        for sn in wb.sheetnames:
            ws = wb[sn]
            rows = []
            for row in ws.iter_rows(values_only=True):
                rows.append(list(row[:12]) if row else [])
            yield sn, rows
    finally:
        wb.close()


def iter_sheet_rows_xlsb(path: Path):
    from pyxlsb import open_workbook

    with open_workbook(path) as wb:
        for sn in wb.sheets:
            rows = []
            with wb.get_sheet(sn) as sheet:
                for row in sheet.rows():
                    vals = [(c.v if c else None) for c in row[:12]]
                    rows.append(vals)
            yield sn, rows


def iter_sheet_rows_xls(path: Path):
    import xlrd

    book = xlrd.open_workbook(path)
    for sn in book.sheet_names():
        sh = book.sheet_by_name(sn)
        rows = []
        for r in range(sh.nrows):
            rows.append([sh.cell_value(r, c) for c in range(min(12, sh.ncols))])
        yield sn, rows


def iter_workbook(path: Path):
    suf = path.suffix.lower()
    if suf == ".xlsx":
        yield from iter_sheet_rows_xlsx(path)
    elif suf == ".xlsb":
        yield from iter_sheet_rows_xlsb(path)
    elif suf == ".xls":
        yield from iter_sheet_rows_xls(path)
    else:
        return


def extract_labels_from_rows(rows: list[list]) -> list[str]:
    labels = []
    for row in rows:
        if not row:
            continue
        # Collect candidate text cells from first 3 columns (BAM puts labels in col B)
        texts = []
        for cell in row[:4]:
            if cell is None:
                continue
            if isinstance(cell, str) and cell.strip():
                texts.append(cell.strip())
            elif isinstance(cell, (int, float)) and not isinstance(cell, bool):
                # skip pure numbers as labels
                continue
        if not texts:
            continue
        # Prefer the longest non-noise text-looking cell
        has_numeric = any(is_numbery(c) for c in row[1:8])
        for raw in texts:
            if len(raw) < 2 or len(raw) > 120:
                continue
            if NOISE_RE.match(raw):
                continue
            if SKIP_LABEL_RE.match(raw.strip()):
                continue
            if GL_CODE_RE.match(raw.strip()):
                # Keep totals only: "Total - 1110 ..." etc. already handled; skip GL lines
                if not raw.lower().startswith("total"):
                    continue
            # Keep section headers even without numbers (Current Assets, etc.)
            # but prefer lines that sit near numbers
            labels.append(raw)
            if has_numeric:
                break
    return labels


def workbook_config_key(cfg: dict) -> str:
    """Cache key component: label extraction depends on the sheet config."""
    return hashlib.sha256(
        json.dumps(
            {
                "sheet_keywords": cfg.get("sheet_keywords") or {},
                "skip_sheets": cfg.get("skip_sheets") or [],
                "strict": cfg["slug"] == "brains-and-motion",
                "v": 1,
            },
            sort_keys=True,
        ).encode("utf-8")
    ).hexdigest()[:12]


def _workbook_worker(job: dict) -> dict:
    """Extract labels from one workbook in a worker process (cached on disk)."""
    doc_extract.quiet_parser_noise()
    cfg = job["cfg"]
    path = Path(job["path"])
    out = {"slug": cfg["slug"], "path": job["path"], "error": None, "cached": False, "labels": 0}
    try:
        key = doc_extract.fingerprint(path, extra=job["config_key"])
    except OSError as exc:
        out["error"] = str(exc)[:300]
        return out
    cached = doc_extract.json_cache_get(key)
    if cached is None:
        by_stmt: dict[str, list[str]] = {"is": [], "bs": [], "cf": [], "other": []}
        try:
            for sn, rows in iter_workbook(path):
                kind = classify_sheet(sn, cfg)
                if kind is None and cfg["slug"] == "brains-and-motion":
                    continue
                if kind is None:
                    # Still scan unknown sheets lightly for well-known totals
                    kind = "other"
                by_stmt[kind].extend(extract_labels_from_rows(rows))
        except Exception as exc:  # noqa: BLE001 - one bad workbook must not stop the run
            out["error"] = f"{type(exc).__name__}: {exc}"[:300]
            return out
        doc_extract.json_cache_put(key, by_stmt)
    else:
        out["cached"] = True
        by_stmt = cached
    out["by_stmt"] = by_stmt
    out["labels"] = sum(len(v) for v in by_stmt.values())
    return out


def extract_workbooks(
    jobs: list[dict],
    *,
    max_workers: int,
    on_done=None,
) -> tuple[dict[tuple[str, str], dict[str, list[str]]], list[dict]]:
    """Extract labels from every company workbook in parallel."""
    results: dict[tuple[str, str], dict[str, list[str]]] = {}
    errors: list[dict] = []
    if not jobs:
        return results, errors
    workers = max(1, min(max_workers, len(jobs)))
    done = 0
    with ProcessPoolExecutor(max_workers=workers) as pool:
        futures = {pool.submit(_workbook_worker, job): job for job in jobs}
        for fut in as_completed(futures):
            job = futures[fut]
            done += 1
            try:
                res = fut.result()
            except Exception as exc:  # noqa: BLE001 - worker crash
                res = {"slug": job["cfg"]["slug"], "path": job["path"], "error": f"worker: {exc}"[:300]}
            if res.get("error"):
                errors.append({"file": res["path"], "error": res["error"]})
            else:
                results[(res["slug"], res["path"])] = res["by_stmt"]
            if on_done:
                on_done(done, len(jobs), res)
    return results, errors


def extract_company_labels(
    cfg: dict,
    *,
    workbook_labels: dict[tuple[str, str], dict[str, list[str]]],
    sources: list[Path] | None = None,
) -> tuple[list[Path], dict[str, list[str]], list[str]]:
    if sources is None:
        sources = pick_source_files(cfg, list_company_files(cfg))
    by_stmt: dict[str, list[str]] = {"is": [], "bs": [], "cf": [], "other": []}
    all_labels: list[str] = []
    for path in sources:
        per_file = workbook_labels.get((cfg["slug"], str(path)))
        if not per_file:
            continue
        for kind in ("is", "bs", "cf", "other"):
            labs = per_file.get(kind) or []
            by_stmt[kind].extend(labs)
            all_labels.extend(labs)

    # unique preserve order
    def uniq(seq):
        out, seen = [], set()
        for x in seq:
            k = canon_label(x)
            if not k or k in seen:
                continue
            seen.add(k)
            out.append(x)
        return out

    by_stmt = {k: uniq(v) for k, v in by_stmt.items()}
    return sources, by_stmt, uniq(all_labels)


def build_core_label_index(concepts: dict) -> dict[str, str]:
    """Map canon label → concept id for mappable core concepts only."""
    idx = {}
    for name, c in concepts.items():
        if not is_mappable_concept(name, c):
            continue
        lab = c.get("l") or ""
        if lab:
            idx[canon_label(lab)] = name
        # local-name fallback for GAAP ids
        if ":" not in name:
            idx[canon_label(name)] = name
    return idx


def is_mappable_concept(name: str, c: dict) -> bool:
    if "Placeholder" in name:
        return False
    if name.startswith("root:") or name.startswith("core:"):
        return False
    if name.startswith("et:") or name.startswith("ea:concept:") or name.startswith("discovery:"):
        return False
    if name.startswith("cdi:") or name.startswith("fibo:"):
        return False
    if name.startswith("nongaap:category:"):
        return False
    if (c.get("a") or c.get("k") == "abstract" or name.endswith("Abstract")) and c.get(
        "layer"
    ) != "standard":
        # XBRL presentation headers and rule containers hold no value and sit
        # outside the browse forest, so evidence must land on a concrete concept.
        return False
    if name.startswith("nongaap:"):
        return "Placeholder" not in name and name not in {
            "nongaap:ARR",
            "nongaap:NetRevenueRetention",
            "nongaap:BeginningRecurringRevenuePlaceholder",
            "nongaap:EndingRecurringRevenuePlaceholder",
        }
    if name.startswith("ea:metric:"):
        return True
    if c.get("layer") == "standard" and c.get("stdKind") in {
        "signal",
        "standardClass",
    }:
        # Curated standard classes and domain signals can be populated from
        # board books / investor memos. Scraped classes are reference-only
        # unless promoted into a curated key class.
        return c.get("category") != "scraped-class"
    return c.get("coreTier") == "core" and c.get("layer") in {"gaap", "core", "nongaap"}


def is_llm_candidate(name: str, c: dict) -> bool:
    """Keep LLM context compact; curated classes map through exact mentions."""
    if not is_mappable_concept(name, c):
        return False
    return c.get("stdKind") != "standardClass"


def exact_label_mappings(
    labels_by_statement: dict[str, list[str]],
    concepts: dict,
) -> tuple[dict[str, str], list[dict], dict[str, list[str]]]:
    """Map exact ontology labels locally and return remaining labels for the LLM."""
    by_label: dict[str, list[str]] = defaultdict(list)
    for cid, concept in concepts.items():
        if not is_mappable_concept(cid, concept):
            continue
        label = canon_label(str(concept.get("l") or ""))
        if label:
            by_label[label].append(cid)
        if ":" not in cid:
            by_label[canon_label(cid)].append(cid)

    def rank(cid: str, statement: str) -> tuple:
        c = concepts[cid]
        if statement in {"is", "bs", "cf"}:
            layer_rank = {"gaap": 0, "nongaap": 1, "standard": 2}.get(c.get("layer"), 3)
        else:
            kind_rank = {"signal": 0, "standardClass": 1}.get(c.get("stdKind"), 2)
            layer_rank = 0 if c.get("layer") == "standard" else 1
            return (kind_rank, layer_rank, cid)
        return (layer_rank, cid)

    mapped: dict[str, str] = {}
    details: list[dict] = []
    remaining: dict[str, list[str]] = {}
    for statement, labels in labels_by_statement.items():
        remaining[statement] = []
        for label in labels:
            candidates = by_label.get(canon_label(label)) or []
            if not candidates:
                remaining[statement].append(label)
                continue
            cid = sorted(set(candidates), key=lambda x: rank(x, statement))[0]
            mapped.setdefault(cid, label)
            details.append(
                {
                    "label": label,
                    "concept": cid,
                    "method": "exact",
                    "score": 1.0,
                    "reason": "Exact canonical label match",
                    "statement": statement,
                }
            )
    return mapped, details, remaining


def is_structural(name: str, c: dict) -> bool:
    if name.startswith("root:"):
        return True
    if name.startswith("core:v12:"):
        return True
    if name.startswith("core:v13:"):
        return True
    if c.get("stdKind") in {"signalDomain", "standard", "scrapeFolder"}:
        return True
    if c.get("extensionPoint"):
        return True
    return False


def attach_parents_catalog(concepts: dict, populated: set[str]) -> list[dict]:
    """Parents the LLM may attach gap concepts under."""
    preferred = [
        "Revenues",
        "CostOfRevenue",
        "GrossProfit",
        "OperatingExpenses",
        "OperatingIncomeLoss",
        "SellingGeneralAndAdministrativeExpense",
        "ResearchAndDevelopmentExpense",
        "InterestAndDebtExpense",
        "NonoperatingIncomeExpense",
        "IncomeTaxExpenseBenefit",
        "ProfitLoss",
        "Assets",
        "AssetsCurrent",
        "AssetsNoncurrent",
        "CashAndCashEquivalentsAtCarryingValue",
        "ReceivablesNetCurrent",
        "InventoryNet",
        "PropertyPlantAndEquipmentNet",
        "IntangibleAssetsNetIncludingGoodwill",
        "Liabilities",
        "LiabilitiesCurrent",
        "LiabilitiesNoncurrent",
        "AccountsPayableAndAccruedLiabilitiesCurrent",
        "ContractWithCustomerLiability",
        "ShortTermBorrowings",
        "LongTermDebt",
        "StockholdersEquity",
        "RetainedEarningsAccumulatedDeficit",
        "NetCashProvidedByUsedInOperatingActivities",
        "NetCashProvidedByUsedInInvestingActivities",
        "NetCashProvidedByUsedInFinancingActivities",
        "core:v12:is_extensions",
        "core:v12:assets.extensions",
        "core:v12:liabilities.extensions",
        "core:v12:equity.extensions",
        "core:v12:cf_extensions",
        "core:v12:metrics_extensions",
        "core:v12:pl_components",
    ]
    out = []
    seen = set()
    for name in preferred:
        if name not in concepts or name in seen:
            continue
        seen.add(name)
        c = concepts[name]
        out.append(
            {
                "id": name,
                "label": c.get("l") or name,
                "populated": name in populated,
                "extensionPoint": bool(c.get("extensionPoint")),
            }
        )
    # Also allow any other currently populated GAAP/nongaap node as parent, as long
    # as it is visible in the browse forest — attaching a gap under an invisible
    # parent would hide the gap node too.
    visible = collect_reachable([r for r in concepts if r.startswith("root:")], concepts)
    for name in sorted(populated):
        if name in seen or name not in concepts:
            continue
        if name.startswith("company:") or name not in visible:
            continue
        c = concepts[name]
        if not is_mappable_concept(name, c) and not name.startswith("core:v12:"):
            continue
        seen.add(name)
        out.append(
            {
                "id": name,
                "label": c.get("l") or name,
                "populated": True,
                "extensionPoint": bool(c.get("extensionPoint")),
            }
        )
    return out


def apply_company_gaps(
    concepts: dict,
    gaps: list[dict],
    slug: str,
    populated: set[str],
    mapped_labels: dict[str, str],
) -> list[str]:
    """Create company:* nodes from LLM gap proposals and wire relationships."""
    created = []
    order = defaultdict(lambda: 1000)
    for g in gaps:
        parent = g.get("parent")
        if not parent or parent not in concepts:
            continue
        cid = f"company:{slug}:{g['id_slug']}"
        if cid in concepts:
            # already present — still ensure links
            pass
        else:
            rel = g.get("relationship") or "child"
            weight = g.get("weight")
            if weight is None and rel == "component":
                weight = 1.0
            try:
                weight_f = float(weight) if weight is not None else None
            except (TypeError, ValueError):
                weight_f = 1.0 if rel == "component" else None

            concepts[cid] = {
                "n": cid,
                "l": g.get("label") or g.get("source_label") or cid,
                "d": g.get("definition")
                or f"Company-specific concept from {slug} financial statements.",
                "t": "company",
                "a": False,
                "p": "",
                "b": "",
                "k": "derived" if rel == "metric" else "atomic",
                "layer": "company",
                "coreTier": "extension",
                "companyGap": True,
                "companyPopulated": True,
                "companyLabels": [g.get("source_label") or g.get("label")],
                "expression": g.get("formula_expression") or "",
                "f": {
                    "atomic": rel != "metric",
                    "combination": False,
                    "classParent": False,
                    "calcTotal": False,
                    "dimensional": False,
                    "ratio": False,
                    "aggregate": False,
                },
                "pc": [],
                "tc": [],
                "cc": [],
                "cp": [],
                "sc": [],
                "gapMeta": {
                    "statement": g.get("statement"),
                    "relationship": rel,
                    "parent": parent,
                    "reason": g.get("reason"),
                    "confidence": g.get("confidence"),
                },
            }
            created.append(cid)

        # Wire browse + formula edges
        rel = g.get("relationship") or "child"
        weight = g.get("weight")
        if weight is None and rel == "component":
            weight = 1.0
        try:
            weight_f = float(weight) if weight is not None else None
        except (TypeError, ValueError):
            weight_f = 1.0 if rel == "component" else None

        parent_c = concepts[parent]
        kids = parent_c.setdefault("tc", [])
        if not any(x.get("c") == cid for x in kids):
            link = {"c": cid, "o": order[parent], "net": "CompanyGap"}
            order[parent] += 10
            if weight_f is not None and rel == "component":
                link["w"] = weight_f
            kids.append(link)

        if rel == "component" and weight_f is not None:
            cc = parent_c.setdefault("cc", [])
            if not any(x.get("c") == cid for x in cc):
                cc.append({"c": cid, "w": weight_f, "net": "CompanyGap"})
            cp = concepts[cid].setdefault("cp", [])
            if not any(x.get("c") == parent for x in cp):
                cp.append({"c": parent, "w": weight_f, "net": "CompanyGap"})

        mapped_labels[cid] = g.get("source_label") or g.get("label") or cid
        populated.add(cid)
        # Parent should remain reachable
        populated.add(parent)

    return created


def available(concept: str, populated: set[str], aliases: dict[str, list[str]]) -> bool:
    if concept in populated:
        return True
    for alt in aliases.get(concept, []):
        if alt in populated:
            return True
    return False


def formula_inputs_for(name: str, concepts: dict) -> list[str] | None:
    """Return required inputs for a concept, or None if not a closed formula."""
    if "Placeholder" in name:
        return None
    if name in {
        "nongaap:ARR",
        "nongaap:NetRevenueRetention",
        "nongaap:AdjustedEBITDA",  # needs non-recurring placeholder — skip auto-derive
    }:
        return None
    if name in EXPLICIT_FORMULAS:
        return list(EXPLICIT_FORMULAS[name])
    c = concepts.get(name) or {}
    df = c.get("df")
    if df:
        # Ignore formulas that depend on placeholders
        if any("Placeholder" in x for x in df):
            return None
        return list(df)
    # Weighted browse children ⇒ formula (e.g. GrossProfit tc)
    tc = c.get("tc") or []
    weighted = [x for x in tc if "w" in x]
    if weighted and 1 < len(weighted) <= 6:
        kids = [x["c"] for x in weighted]
        if any("Placeholder" in k for k in kids):
            return None
        return kids
    return None


def close_formulas(seed: set[str], concepts: dict, aliases: dict[str, list[str]]) -> set[str]:
    populated = set(seed)
    # Candidates: anything with df / explicit formula / weighted tc, plus nongaap/metrics
    candidates = set(EXPLICIT_FORMULAS)
    for name, c in concepts.items():
        if c.get("df"):
            candidates.add(name)
        tc = c.get("tc") or []
        if any("w" in x for x in tc) and 1 < len([x for x in tc if "w" in x]) <= 6:
            candidates.add(name)
    changed = True
    while changed:
        changed = False
        for name in list(candidates):
            if name in populated:
                continue
            inputs = formula_inputs_for(name, concepts)
            if not inputs:
                continue
            if all(available(inp, populated, aliases) for inp in inputs):
                populated.add(name)
                changed = True
    return populated


def browse_children(c: dict) -> list[dict]:
    return list(c.get("tc") or [])


def collect_reachable(roots: list[str], concepts: dict) -> set[str]:
    seen = set()
    stack = list(roots)
    while stack:
        n = stack.pop()
        if n in seen or n not in concepts:
            continue
        seen.add(n)
        for link in browse_children(concepts[n]):
            stack.append(link["c"])
    return seen


def prune_forest(
    roots: list[str],
    concepts: dict,
    populated: set[str],
) -> tuple[list[str], set[str]]:
    """Keep populateable nodes + structural ancestors. Drop empty shells."""

    memo: dict[str, bool] = {}

    def decide(name: str) -> bool:
        if name in memo:
            return memo[name]
        if name not in concepts:
            memo[name] = False
            return False
        # Drop Guidance / Reference for company trees
        if name in {"root:Guidance", "root:Reference"}:
            memo[name] = False
            return False
        c = concepts[name]
        raw_kids = browse_children(c)
        kept_links = []
        for link in raw_kids:
            child = link["c"]
            if decide(child):
                kept_links.append(link)
        c["_kept_tc"] = kept_links

        if name in populated:
            ok = True
        elif is_structural(name, c) and kept_links:
            ok = True
        else:
            # Non-structural (metrics, GAAP lines): only if populateable
            ok = False

        # Empty extension hooks go away
        if (c.get("extensionPoint") or str(name).endswith(".extensions")) and not kept_links:
            ok = False

        memo[name] = ok
        return ok

    new_roots = [r for r in roots if decide(r)]
    kept_nodes = {n for n, ok in memo.items() if ok}
    return new_roots, kept_nodes


EXTENSION_HOOKS = {
    "assets": "core:v12:assets.extensions",
    "liabilities": "core:v12:liabilities.extensions",
    "equity": "core:v12:equity.extensions",
    "is": "core:v12:is_extensions",
    "cf": "core:v12:cf_extensions",
    "metrics": "core:v12:metrics_extensions",
}

ASSET_PREFIXES = (
    "Assets",
    "Cash",
    "Receivables",
    "AccountsReceivable",
    "Inventory",
    "PrepaidExpense",
    "PropertyPlantAndEquipment",
    "IntangibleAssets",
    "Goodwill",
    "OperatingLeaseRightOfUseAsset",
    "ContractWithCustomerAsset",
    "LongTermInvestments",
    "ShortTermInvestments",
    "DeferredTaxAssets",
)
LIAB_PREFIXES = (
    "Liabilit",
    "AccountsPayable",
    "ShortTermBorrowings",
    "LongTermDebt",
    "ContractWithCustomerLiability",
    "OperatingLeaseLiability",
    "DeferredIncomeTax",
    "NotesPayable",
    "ConvertibleNotesPayable",
)


def extension_bucket(name: str) -> str | None:
    """Which statement extension hook a concept belongs under, if any."""
    stem = name[: -len("Abstract")] if name.endswith("Abstract") else name
    if stem in ASSET_CONCEPTS or stem.startswith(ASSET_PREFIXES):
        return "assets"
    if stem in LIAB_CONCEPTS or stem.startswith(LIAB_PREFIXES):
        return "liabilities"
    if stem in EQUITY_CONCEPTS or "Equity" in stem or stem.endswith("StockValue"):
        return "equity"
    if stem in IS_CONCEPTS or stem in {
        "nongaap:EBITDA",
        "nongaap:GrossMargin",
        "nongaap:OperatingMargin",
    }:
        return "is"
    if stem in CF_CONCEPTS or stem.startswith("NetCash") or stem.startswith("PaymentsTo"):
        return "cf"
    if stem.startswith("nongaap:") or stem.startswith("ea:metric:"):
        return "metrics"
    return None


def attach_extensions(
    concepts: dict,
    populated: set[str],
    kept: set[str],
    mapped_labels: dict[str, str],
) -> None:
    """Put mapped concepts not already in the browse forest under extension hooks."""
    hooks = EXTENSION_HOOKS
    # Ensure hooks exist in kept set if we attach anything
    reachable = collect_reachable(
        [r for r in concepts if r.startswith("root:")], concepts
    )
    bucket = extension_bucket

    order_counters = defaultdict(lambda: 10)
    for name in sorted(populated):
        if name in reachable and name in kept:
            continue
        b = bucket(name)
        if not b:
            continue
        hook = hooks[b]
        if hook not in concepts:
            continue
        # Skip if already a direct child of something kept
        already = False
        for n in kept:
            for link in browse_children(concepts.get(n) or {}):
                if link["c"] == name:
                    already = True
                    break
            if already:
                break
        if already:
            continue
        kids = concepts[hook].setdefault("tc", [])
        if any(x.get("c") == name for x in kids):
            continue
        kids.append({"c": name, "o": order_counters[hook], "net": "CompanyExtension"})
        order_counters[hook] += 10
        # ensure hook kept / parent chain
        concepts[hook]["extensionPoint"] = True
        # annotate source label
        if name in concepts and name in mapped_labels:
            concepts[name].setdefault("companyLabels", [])
            if mapped_labels[name] not in concepts[name]["companyLabels"]:
                concepts[name]["companyLabels"].append(mapped_labels[name])
        kept.add(name)
        kept.add(hook)


def rehome_orphans(
    work: dict,
    base_concepts: dict,
    populated: set[str],
    kept: set[str],
    roots: list[str],
) -> tuple[list[dict], list[str]]:
    """Re-home kept concepts that are unreachable from any browse root.

    Some populated concepts only hang off curated parents that are not part of the
    browse forest (the XBRL ``*Abstract`` skeleton, or concepts the forest omits).
    Keeping them regardless left them in the payload but invisible in the tree, so
    node counts disagreed with what the viewer could show. Attach them to the
    nearest visible ancestor, else the right extension hook, else drop them.
    """
    reachable = collect_reachable(roots, work)
    parents_of: dict[str, list[str]] = defaultdict(list)
    for pid, pc in base_concepts.items():
        for link in browse_children(pc):
            parents_of[link["c"]].append(pid)
    # Company gap nodes record their intended parent as a formula parent.
    for name, c in work.items():
        if name.startswith("company:"):
            for link in c.get("cp") or []:
                if isinstance(link, dict) and link.get("c"):
                    parents_of[name].append(link["c"])

    def nearest_visible(name: str) -> str | None:
        seen = {name}
        frontier = list(parents_of.get(name) or [])
        for _ in range(8):
            nxt = []
            for cand in frontier:
                if cand in seen:
                    continue
                seen.add(cand)
                if cand in reachable and cand in work:
                    return cand
                nxt.extend(parents_of.get(cand) or [])
            if not nxt:
                break
            frontier = nxt
        return None

    attached: list[dict] = []
    dropped: list[str] = []
    order_counters = defaultdict(lambda: 900)

    def attach(parent: str, child: str) -> None:
        kids = work[parent].setdefault("tc", [])
        if not any(x.get("c") == child for x in kids):
            kids.append({"c": child, "o": order_counters[parent], "net": "CompanyExtension"})
            order_counters[parent] += 10
        kept.add(child)
        reachable.add(child)

    def hook_for(name: str) -> str | None:
        bucket = extension_bucket(name)
        hook = EXTENSION_HOOKS.get(bucket) if bucket else None
        if not hook or hook not in work:
            return None
        if hook not in reachable:
            # Pruning drops empty hooks; revive one if its section is still visible.
            hook_host = nearest_visible(hook)
            if not hook_host:
                return None
            work[hook]["extensionPoint"] = True
            attach(hook_host, hook)
        return hook

    pending = sorted(((kept | populated) & set(work)) - reachable)
    # Attaching one orphan can make another attachable, so iterate to a fixpoint.
    for _ in range(3):
        if not pending:
            break
        still: list[str] = []
        for name in pending:
            host, via = nearest_visible(name), "ancestor"
            if not host:
                host, via = hook_for(name), "hook"
            if not host:
                still.append(name)
                continue
            attach(host, name)
            attached.append({"concept": name, "parent": host, "via": via})
        if len(still) == len(pending):
            break
        pending = still
    for name in pending:
        kept.discard(name)
        dropped.append(name)
    return attached, dropped


def apply_kept_tc(concepts: dict, kept: set[str]) -> None:
    for name, c in concepts.items():
        if "_kept_tc" in c:
            c["tc"] = [x for x in c["_kept_tc"] if x["c"] in kept]
            del c["_kept_tc"]
        elif name in kept:
            c["tc"] = [x for x in (c.get("tc") or []) if x["c"] in kept]
        # Filter calc / presentation / used-by to kept only for cleaner Formulas mode
        if name in kept:
            for key in ("cc", "pc", "cp"):
                if key in c and isinstance(c[key], list):
                    c[key] = [x for x in c[key] if isinstance(x, dict) and x.get("c") in kept]
            if "sc" in c and isinstance(c["sc"], list):
                c["sc"] = [x for x in c["sc"] if x in kept]
            if "df" in c and isinstance(c["df"], list):
                # keep df for transparency even if some missing — but trim to available
                pass


def slim_payload(
    base: dict,
    kept: set[str],
    roots: list[str],
    meta_extra: dict,
    mappings: list[dict],
    gaps: list[dict] | None = None,
) -> dict:
    concepts = {k: v for k, v in base["concepts"].items() if k in kept}
    for c in concepts.values():
        c.pop("_kept_tc", None)

    summary = dict(base.get("summary") or {})
    summary.update(
        {
            "ontologyNodes": len(concepts),
            "ontologyRoots": len(roots),
            "forestRoots": len(roots),
            "companyMappedLabels": sum(1 for m in mappings if m.get("concept")),
            "companyUnmappedLabels": sum(1 for m in mappings if not m.get("concept")),
            "companyPopulatedConcepts": meta_extra.get("populatedCount", 0),
            "companyGapConcepts": meta_extra.get("gapCount", 0),
        }
    )
    meta = dict(base.get("meta") or {})
    meta.update(meta_extra)
    meta["companyTree"] = True
    meta["defaultBrowseMode"] = "presentation"

    return {
        "meta": meta,
        "summary": summary,
        "ontologyRoots": roots,
        "classRoots": [r for r in (base.get("classRoots") or []) if r in kept],
        "presentationNetworkRoots": base.get("presentationNetworkRoots") or {},
        "browseModes": base.get("browseModes"),
        "concepts": concepts,
        "companyMappings": [m for m in mappings if m.get("concept")][:500],
        "companyGaps": gaps or [],
    }


def now_iso() -> str:
    return datetime.now(timezone.utc).isoformat(timespec="seconds")


class RunProgress:
    """Thread-safe run state, mirrored to company_trees/build_progress.json."""

    def __init__(self, path: Path, run_id: str, slugs: list[str]):
        self.path = path
        self.started = time.time()
        self.lock = threading.Lock()
        self.last_write = 0.0
        self.state: dict = {
            "runId": run_id,
            "startedAt": now_iso(),
            "phase": "starting",
            "extraction": {},
            "companies": {slug: {"stage": "queued"} for slug in slugs},
        }
        self._write(force=True)

    def _write(self, *, force: bool) -> None:
        """Best-effort mirror of run state; never raises into the build."""
        now = time.time()
        if not force and now - self.last_write < 0.5:
            return
        self.last_write = now
        self.state["elapsedSec"] = round(now - self.started, 1)
        self.state["updatedAt"] = now_iso()
        payload = json.dumps(self.state, indent=2)
        self.path.parent.mkdir(parents=True, exist_ok=True)
        tmp = self.path.with_name(f"{self.path.stem}.{os.getpid()}.tmp")
        for attempt in range(3):
            try:
                tmp.write_text(payload, encoding="utf-8")
                tmp.replace(self.path)
                return
            except OSError:
                # Windows hands out transient sharing violations on rapid replaces.
                time.sleep(0.05 * (attempt + 1))
        try:
            self.path.write_text(payload, encoding="utf-8")
        except OSError as exc:
            log.debug("progress write skipped: %s", exc)

    def phase(self, name: str, **fields) -> None:
        with self.lock:
            self.state["phase"] = name
            self.state.update(fields)
            self._write(force=True)

    def extraction(self, **fields) -> None:
        with self.lock:
            self.state["extraction"].update(fields)
            self._write(force=False)

    def company(self, slug: str, **fields) -> None:
        with self.lock:
            entry = self.state["companies"].setdefault(slug, {})
            entry.update(fields)
            entry["atSec"] = round(time.time() - self.started, 1)
            self._write(force=fields.get("stage") in {"done", "failed", "starting"})


def setup_logging(run_id: str, verbose: bool = False) -> Path:
    """Line-flushed console logging plus a full debug log file."""
    LOG_DIR.mkdir(parents=True, exist_ok=True)
    log_path = LOG_DIR / f"company_build_{run_id}.log"
    fmt = logging.Formatter(
        "%(asctime)s %(levelname).1s %(name)-7s %(message)s", datefmt="%H:%M:%S"
    )
    root = logging.getLogger()
    root.setLevel(logging.DEBUG)
    for handler in list(root.handlers):
        root.removeHandler(handler)
    try:
        sys.stdout.reconfigure(line_buffering=True)
    except Exception:  # noqa: BLE001 - older/odd stdout objects
        pass
    stream = logging.StreamHandler(sys.stdout)
    stream.setLevel(logging.DEBUG if verbose else logging.INFO)
    stream.setFormatter(fmt)
    file_handler = logging.FileHandler(log_path, encoding="utf-8")
    file_handler.setLevel(logging.DEBUG)
    file_handler.setFormatter(fmt)
    root.addHandler(stream)
    root.addHandler(file_handler)
    for noisy in ("httpx", "httpcore", "anthropic", "urllib3"):
        logging.getLogger(noisy).setLevel(logging.WARNING)
    doc_extract.quiet_parser_noise()
    return log_path


def prepare_sources(
    selected: list[dict],
    *,
    max_workers: int,
    progress: RunProgress | None = None,
) -> dict[str, dict]:
    """Phase 1: list and extract every source document for every company in parallel."""
    plans: dict[str, dict] = {}
    workbook_jobs: list[dict] = []
    narrative_paths: list[Path] = []
    for cfg in selected:
        financial = pick_source_files(cfg, list_company_files(cfg))
        narrative = list_narrative_files(cfg)
        plans[cfg["slug"]] = {
            "financial_sources": financial,
            "narrative_sources": narrative,
        }
        config_key = workbook_config_key(cfg)
        workbook_jobs.extend(
            {"cfg": cfg, "path": str(p), "config_key": config_key} for p in financial
        )
        narrative_paths.extend(narrative)
        log.info(
            "[%s] sources: %d workbooks, %d documents",
            cfg["name"],
            len(financial),
            len(narrative),
        )

    total = len(workbook_jobs) + len(narrative_paths)
    log.info(
        "extracting %d files with %d worker processes (cached results are reused)",
        total,
        max_workers,
    )
    if progress:
        progress.extraction(
            totalFiles=total,
            workbooks=len(workbook_jobs),
            documents=len(narrative_paths),
            done=0,
        )

    counters = {"done": 0, "cached": 0, "errors": 0}
    started = time.time()

    def on_done(done: int, total_files: int, res: dict) -> None:
        counters["done"] += 1
        if res.get("cached"):
            counters["cached"] += 1
        if res.get("error"):
            counters["errors"] += 1
            log.warning("extract failed: %s (%s)", Path(res["path"]).name, res["error"])
        n = counters["done"]
        if n % 5 == 0 or n == total:
            rate = n / max(0.1, time.time() - started)
            log.info(
                "extract %d/%d (%d cached, %d errors, %.1f files/s)",
                n,
                total,
                counters["cached"],
                counters["errors"],
                rate,
            )
        if progress:
            progress.extraction(
                done=n, cached=counters["cached"], errors=counters["errors"]
            )

    workbook_labels, workbook_errors = extract_workbooks(
        workbook_jobs, max_workers=max_workers, on_done=on_done
    )
    texts, text_errors = doc_extract.extract_texts(
        narrative_paths, max_workers=max_workers, on_done=on_done
    )
    log.info(
        "extraction done in %.1fs (%d files, %d cached, %d errors)",
        time.time() - started,
        counters["done"],
        counters["cached"],
        counters["errors"],
    )

    errors_by_slug: dict[str, list[dict]] = defaultdict(list)
    for err in workbook_errors + text_errors:
        rel = err["file"]
        for cfg in selected:
            folder = str(DOCUMENT_PACK_DIR / cfg["folder"])
            if rel.startswith(folder):
                errors_by_slug[cfg["slug"]].append(
                    {"file": Path(rel).name, "error": err["error"]}
                )
                break

    prepared: dict[str, dict] = {}
    for cfg in selected:
        slug = cfg["slug"]
        plan = plans[slug]
        prepared[slug] = {
            **plan,
            "workbook_labels": workbook_labels,
            "texts": texts,
            "extraction_errors": errors_by_slug.get(slug, []),
        }
    return prepared


def build_company_tree(
    cfg: dict,
    base: dict,
    aliases: dict,
    *,
    prepared: dict,
    refresh_llm: bool = False,
    llm_workers: int = 6,
    progress: "RunProgress | None" = None,
) -> dict:
    slug = cfg["slug"]
    started = time.time()
    timings: dict[str, float] = {}
    last = [started]

    def stage(name: str, **fields) -> None:
        now = time.time()
        timings[name] = round(now - last[0], 1)
        last[0] = now
        log.info("[%s] %s (%.1fs, %.1fs total)", cfg["name"], name, timings[name], now - started)
        if progress:
            progress.company(slug, stage=name, **fields)

    concepts = base["concepts"]
    financial_sources = prepared["financial_sources"]
    narrative_sources = prepared["narrative_sources"]
    extraction_errors = list(prepared["extraction_errors"])

    _, by_stmt, labels = extract_company_labels(
        cfg, workbook_labels=prepared["workbook_labels"], sources=financial_sources
    )
    stage(
        "workbook-labels",
        workbookLabels=len(labels),
        financialSources=len(financial_sources),
    )
    _, narrative_labels, narrative_exact = extract_narrative_evidence(
        cfg, concepts, texts=prepared["texts"], sources=narrative_sources
    )
    stage(
        "document-mining",
        documentLines=len(narrative_labels),
        documentMentions=len(narrative_exact),
        narrativeSources=len(narrative_sources),
    )
    by_stmt["document"] = narrative_labels
    sources = financial_sources + narrative_sources

    exact_mapped, exact_details, llm_labels = exact_label_mappings(by_stmt, concepts)
    stage(
        "exact-mapping",
        exactMapped=len(exact_mapped),
        llmQueueLabels=sum(len(v) for v in llm_labels.values()),
    )

    candidates = build_candidate_list(concepts, is_llm_candidate)
    cache_dir = OUT_DIR / cfg["slug"]
    llm_mapped, llm_details = map_labels_with_llm(
        company_name=cfg["name"],
        labels_by_statement=llm_labels,
        candidates=candidates,
        cache_path=cache_dir / "llm_map_cache.json",
        refresh=refresh_llm,
        workers=llm_workers,
        progress=(
            (lambda kind, done, total: progress.company(slug, stage=f"llm-{kind}", batches=f"{done}/{total}"))
            if progress
            else None
        ),
    )
    stage("llm-mapping", llmMapped=len(llm_mapped))
    mapped = {**exact_mapped, **llm_mapped, **narrative_exact}
    details = exact_details + llm_details
    for cid, evidence in narrative_exact.items():
        details.append(
            {
                "label": evidence,
                "concept": cid,
                "method": "document-mention",
                "score": 1.0,
                "reason": "Curated ontology phrase found in document corpus",
                "statement": "document",
            }
        )

    # Drop any accidental non-mappable ids
    mapped = {
        k: v
        for k, v in mapped.items()
        if k in concepts and is_mappable_concept(k, concepts[k])
    }

    seed = set(mapped.keys())
    work = json.loads(json.dumps(concepts))
    populated = close_formulas(seed, work, aliases)
    stage("formula-closure", seed=len(seed), populated=len(populated))

    for concept, lab in mapped.items():
        if concept in work:
            work[concept].setdefault("companyLabels", [])
            if lab not in work[concept]["companyLabels"]:
                work[concept]["companyLabels"].append(lab)
            work[concept]["companyPopulated"] = True
    for name in populated:
        if name in work:
            work[name]["companyPopulated"] = True
            if name not in seed:
                work[name]["companyDerived"] = True

    # --- Gap analysis: important unmapped lines → new company nodes ---
    unmapped_rows = [m for m in details if not m.get("concept")]
    mapped_concept_info = [
        {"id": cid, "label": work[cid].get("l", cid), "example": mapped[cid]}
        for cid in sorted(seed)
        if cid in work
    ]
    parents = attach_parents_catalog(work, populated)
    gaps = propose_gaps_with_llm(
        company_name=cfg["name"],
        company_slug=cfg["slug"],
        unmapped_labels=unmapped_rows,
        mapped_concepts=mapped_concept_info,
        attach_parents=parents,
        cache_path=cache_dir / "llm_gap_cache.json",
        refresh=refresh_llm,
        workers=llm_workers,
        progress=(
            (lambda kind, done, total: progress.company(slug, stage=f"llm-{kind}", batches=f"{done}/{total}"))
            if progress
            else None
        ),
    )
    created_gaps = apply_company_gaps(work, gaps, cfg["slug"], populated, mapped)
    # Re-close formulas in case gap metrics introduced df-style nodes (none yet)
    populated = close_formulas(populated, work, aliases)
    stage("gap-analysis", gapProposals=len(gaps), gapsCreated=len(created_gaps))

    roots = list(base.get("ontologyRoots") or [])
    base_work = {**base, "concepts": work}
    new_roots, kept = prune_forest(roots, work, populated)
    attach_extensions(work, populated, kept, mapped)

    for c in work.values():
        c.pop("_kept_tc", None)
    new_roots, kept = prune_forest(roots, work, populated | kept)
    attach_extensions(work, populated, kept, mapped)
    apply_kept_tc(work, kept)
    rehomed, dropped_orphans = rehome_orphans(work, concepts, populated, kept, new_roots)
    if rehomed or dropped_orphans:
        log.info(
            "[%s] re-homed %d unreachable concepts, dropped %d",
            cfg["name"],
            len(rehomed),
            len(dropped_orphans),
        )
    populated &= kept
    stage("prune", nodes=len(kept), roots=len(new_roots))

    meta_extra = {
        "company": cfg["name"],
        "companySlug": cfg["slug"],
        "sourceFiles": [p.name for p in sources],
        "financialSourceCount": len(financial_sources),
        "narrativeSourceCount": len(narrative_sources),
        "sourceDigest": hashlib.sha256(
            "\n".join(
                f"{p.relative_to(ROOT)}:{p.stat().st_size}:{p.stat().st_mtime_ns}"
                for p in sources
            ).encode("utf-8")
        ).hexdigest()[:16],
        "extractionErrorCount": len(extraction_errors),
        "baseOntologyVersion": (base.get("meta") or {}).get("ontologyVersion", "1.31"),
        "mappingMethod": "exact + document mentions + claude-llm",
        "populatedCount": len(populated),
        "seedMappedCount": len(seed),
        "derivedCount": len(populated - seed - set(created_gaps)),
        "gapCount": len(created_gaps),
        "buildSeconds": round(time.time() - started, 1),
        "notes": (
            "Company tree pruned from the base ontology using Claude mapping + gap analysis. "
            "Financial statements, board books, and investor memos are scanned. Existing "
            "financial, strategic, operational, governance, and standards concepts are kept "
            "when evidenced; important unmapped lines are added as company:* nodes."
        ),
    }

    payload = slim_payload(base_work, kept, new_roots, meta_extra, details, gaps=gaps)
    payload["_debug"] = {
        "labelsSample": labels[:40],
        "byStatementCounts": {k: len(v) for k, v in by_stmt.items()},
        "financialSources": [str(p.relative_to(ROOT)) for p in financial_sources],
        "narrativeSources": [str(p.relative_to(ROOT)) for p in narrative_sources],
        "extractionErrors": extraction_errors,
        "documentExactConcepts": sorted(narrative_exact),
        "seed": sorted(seed),
        "gapsCreated": created_gaps,
        "derived": sorted(populated - seed - set(created_gaps)),
        "populated": sorted(populated),
        "unmappedSample": [m["label"] for m in details if not m.get("concept")][:40],
        "rehomedConcepts": rehomed,
        "droppedUnreachable": dropped_orphans,
        "stageSeconds": timings,
    }
    return payload


def write_company(payload: dict, slug: str) -> Path:
    out = OUT_DIR / slug
    out.mkdir(parents=True, exist_ok=True)
    debug = payload.pop("_debug", None)
    path = out / "taxonomy-data.json"
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    if debug:
        (out / "coverage.json").write_text(
            json.dumps(debug, ensure_ascii=False, indent=2), encoding="utf-8"
        )
    if payload.get("companyGaps") is not None:
        (out / "gaps.json").write_text(
            json.dumps(payload["companyGaps"], ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
    vdir = VIEWER_DIR / slug
    vdir.mkdir(parents=True, exist_ok=True)
    shutil.copy2(path, vdir / "taxonomy-data.json")
    # GitHub Pages serves from repo root
    pdir = PAGES_DIR / slug
    pdir.mkdir(parents=True, exist_ok=True)
    shutil.copy2(path, pdir / "taxonomy-data.json")
    return path


def write_run_report(run_id: str, rows: list[dict], failures: list[dict], elapsed: float) -> Path:
    report = {
        "runId": run_id,
        "finishedAt": now_iso(),
        "elapsedSec": round(elapsed, 1),
        "baseOntology": str(ONTOLOGY_PATH.relative_to(ROOT)),
        "companies": rows,
        "failures": failures,
    }
    json_path = OUT_DIR / "run_report.json"
    json_path.write_text(json.dumps(report, indent=2), encoding="utf-8")

    lines = [
        f"# Company tree build {run_id}",
        "",
        f"- base ontology: `{report['baseOntology']}`",
        f"- wall clock: {report['elapsedSec']}s",
        f"- companies built: {len(rows)}" + (f", failed: {len(failures)}" if failures else ""),
        "",
        "| Company | Sources (wb/doc) | Labels | Mapped | Gaps | Derived | Populated | Nodes | Roots | Extract errors | Sec |",
        "| --- | --- | --- | --- | --- | --- | --- | --- | --- | --- | --- |",
    ]
    for r in rows:
        lines.append(
            f"| {r['company']} | {r['financialSources']}/{r['narrativeSources']} | {r['labels']} "
            f"| {r['mapped']} | {r['gaps']} | {r['derived']} | {r['populated']} | {r['nodes']} "
            f"| {r['roots']} | {r['extractionErrors']} | {r['seconds']} |"
        )
    if failures:
        lines += ["", "## Failures", ""]
        lines += [f"- **{f['company']}**: {f['error']}" for f in failures]
    md_path = OUT_DIR / "run_report.md"
    md_path.write_text("\n".join(lines) + "\n", encoding="utf-8")
    return md_path


def main() -> None:
    parser = argparse.ArgumentParser(description="Build company ontology trees via Claude")
    parser.add_argument(
        "--refresh-llm",
        action="store_true",
        help="Ignore LLM caches and re-query Claude",
    )
    parser.add_argument(
        "--company",
        action="append",
        dest="companies",
        help="Only build this company slug (repeatable)",
    )
    parser.add_argument(
        "--extract-workers",
        type=int,
        default=int(os.environ.get("EXTRACT_WORKERS", "0")) or max(2, (os.cpu_count() or 4) - 1),
        help="Worker processes for document/workbook extraction",
    )
    parser.add_argument(
        "--company-workers",
        type=int,
        default=int(os.environ.get("COMPANY_WORKERS", "0")) or 5,
        help="Companies built concurrently",
    )
    parser.add_argument(
        "--llm-workers",
        type=int,
        default=int(os.environ.get("LLM_BATCH_WORKERS", "0")) or 6,
        help="Concurrent LLM batches per company",
    )
    parser.add_argument("--verbose", action="store_true", help="Log per-request LLM detail")
    args = parser.parse_args()

    run_id = datetime.now().strftime("%Y%m%d_%H%M%S")
    log_path = setup_logging(run_id, verbose=args.verbose)

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    VIEWER_DIR.mkdir(parents=True, exist_ok=True)
    PAGES_DIR.mkdir(parents=True, exist_ok=True)

    selected = [c for c in COMPANIES if not args.companies or c["slug"] in args.companies]
    progress = RunProgress(PROGRESS_PATH, run_id, [c["slug"] for c in selected])
    started = time.time()

    log.info("run %s | log %s | progress %s", run_id, log_path.name, PROGRESS_PATH.name)
    log.info(
        "companies: %s | extract workers=%d, company workers=%d, llm batch workers=%d",
        ", ".join(c["slug"] for c in selected),
        args.extract_workers,
        args.company_workers,
        args.llm_workers,
    )

    log.info("loading base ontology %s", ONTOLOGY_PATH.relative_to(ROOT))
    progress.phase("load-ontology")
    base = json.loads(ONTOLOGY_PATH.read_text(encoding="utf-8"))
    _, aliases = load_synonyms()
    log.info("base ontology: %d concepts", len(base["concepts"]))

    progress.phase("extract-sources")
    prepared = prepare_sources(selected, max_workers=args.extract_workers, progress=progress)

    catalog = [
        {
            "id": "v1.31",
            "label": "Ontology v1.31 (full core)",
            "path": "taxonomy-data.json",
        }
    ]
    rows: list[dict] = []
    failures: list[dict] = []
    write_lock = threading.Lock()
    progress.phase("build-companies")

    def build_one(cfg: dict) -> dict | None:
        slug = cfg["slug"]
        try:
            progress.company(slug, stage="starting")
            payload = build_company_tree(
                cfg,
                base,
                aliases,
                prepared=prepared[slug],
                refresh_llm=args.refresh_llm,
                llm_workers=args.llm_workers,
                progress=progress,
            )
            meta = payload["meta"]
            nodes = len(payload["concepts"])
            roots = len(payload["ontologyRoots"])
            debug = payload.get("_debug") or {}
            with write_lock:
                path = write_company(payload, slug)
            row = {
                "company": cfg["name"],
                "slug": slug,
                "financialSources": meta["financialSourceCount"],
                "narrativeSources": meta["narrativeSourceCount"],
                "labels": sum((debug.get("byStatementCounts") or {}).values()),
                "mapped": meta["seedMappedCount"],
                "gaps": meta["gapCount"],
                "derived": meta["derivedCount"],
                "populated": meta["populatedCount"],
                "nodes": nodes,
                "roots": roots,
                "extractionErrors": meta["extractionErrorCount"],
                "seconds": meta["buildSeconds"],
                "output": str(path.relative_to(ROOT)),
                "stageSeconds": debug.get("stageSeconds") or {},
            }
            log.info(
                "[%s] DONE in %.0fs: mapped=%d gaps=%d derived=%d populated=%d nodes=%d roots=%d -> %s",
                cfg["name"],
                row["seconds"],
                row["mapped"],
                row["gaps"],
                row["derived"],
                row["populated"],
                nodes,
                roots,
                row["output"],
            )
            progress.company(slug, stage="done", **{k: row[k] for k in ("mapped", "gaps", "populated", "nodes")})
            return row
        except Exception as exc:  # noqa: BLE001 - one company must not sink the run
            log.exception("[%s] FAILED: %s", cfg["name"], exc)
            progress.company(slug, stage="failed", error=str(exc)[:300])
            with write_lock:
                failures.append({"company": cfg["name"], "slug": slug, "error": str(exc)[:500]})
            return None

    workers = max(1, min(args.company_workers, len(selected)))
    with ThreadPoolExecutor(max_workers=workers) as pool:
        futures = {pool.submit(build_one, cfg): cfg for cfg in selected}
        for fut in as_completed(futures):
            row = fut.result()
            if row:
                with write_lock:
                    rows.append(row)

    order = {c["slug"]: i for i, c in enumerate(COMPANIES)}
    rows.sort(key=lambda r: order.get(r["slug"], 999))
    for row in rows:
        catalog.append(
            {
                "id": row["slug"],
                "label": row["company"],
                "path": f"companies/{row['slug']}/taxonomy-data.json",
                "populated": row["populated"],
                "gaps": row["gaps"],
                "nodes": row["nodes"],
            }
        )

    # If building a subset, merge with existing catalog entries for other companies
    catalog_path = VIEWER_DIR / "catalog.json"
    if args.companies and catalog_path.exists():
        old = json.loads(catalog_path.read_text(encoding="utf-8"))
        by_id = {x["id"]: x for x in old}
        for entry in catalog:
            by_id[entry["id"]] = entry
        # preserve order: base tree then companies
        base_id = next((x for x in ("v1.31", "v1.3", "v1.2") if x in by_id), None)
        ordered = [by_id[base_id]] if base_id else []
        for c in COMPANIES:
            if c["slug"] in by_id:
                ordered.append(by_id[c["slug"]])
        catalog = ordered

    catalog_path.write_text(json.dumps(catalog, indent=2), encoding="utf-8")
    (OUT_DIR / "catalog.json").write_text(json.dumps(catalog, indent=2), encoding="utf-8")
    (PAGES_DIR / "catalog.json").write_text(json.dumps(catalog, indent=2), encoding="utf-8")

    elapsed = time.time() - started
    report = write_run_report(run_id, rows, failures, elapsed)
    progress.phase("done", finishedAt=now_iso())
    log.info("catalog -> %s", catalog_path.relative_to(ROOT))
    log.info("report  -> %s", report.relative_to(ROOT))
    log.info(
        "run %s finished in %.1fs: %d built, %d failed",
        run_id,
        elapsed,
        len(rows),
        len(failures),
    )
    if failures:
        sys.exit(1)


if __name__ == "__main__":
    main()
